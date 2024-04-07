import csv
import zipfile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from tests.conftest import archive


def test_csv():
    with zipfile.ZipFile(archive, 'r') as zip_file:  # открываем архив
        with zip_file.open('example_1.5kb.csv') as csv_file:  # открываем файл в архиве
            content = csv_file.read().decode(
                'utf-8-sig')  # читаем содержимое файла и декодируем его если в файле есть символы не из английского алфавита
            csvreader = list(csv.reader(content.splitlines()))  # читаем содержимое файла и преобразуем его в список
        assert csvreader[0][0] == 'name'


def test_pdf():
    with zipfile.ZipFile(archive, 'r') as zip_file:  # открываем архив
        with zip_file.open('Python Testing with Pytest (Brian Okken).pdf') as pdf_file:
            reader = PdfReader(pdf_file)  # открываем файл
            pdf_content = reader.pages[1].extract_text()
    assert 'Python Testing with pytest' in pdf_content


def test_xlsx():
    with zipfile.ZipFile(archive, 'r') as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
    assert sheet.cell(row=3, column=2).value == 'Mara'
